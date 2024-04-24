import openpyxl
import time

from selenium.webdriver.common.by import By
from selenium import webdriver



def get_cards(driver):
    scroll(30, 1, driver)
    cards = (driver.find_elements(by=By.CLASS_NAME, value="product-card__wrapper"))
    return cards


def scroll(count, delay, driver):
    for i in range(count):
        driver.execute_script("window.scrollBy(0, 400)")
        time.sleep(delay)


def parse(card):
    base = []
    name = (card.find_element(by=By.CLASS_NAME, value='product-card__brand'))
    base.append(f'{name.text}')

    discription = (card.find_element(by=By.CLASS_NAME, value='product-card__name'))
    base.append(f'{discription.text}')

    photo = (card.find_element(by=By.CLASS_NAME, value='j-thumbnail').get_attribute('src'))
    base.append(f'{photo}')

    cost = (card.find_element(by=By.TAG_NAME, value='ins').text)
    base.append(f'{cost}')

    rate = (card.find_element(by=By.CLASS_NAME, value='product-card__rating-wrap').find_elements(by=By.TAG_NAME, value='span'))
    base.append(rate[0].text)
    base.append(rate[1].text)

    return base



def main():
    url =r'https://www.wildberries.ru/catalog/muzhchinam/pizhamy'
    service = webdriver.EdgeService(executable_path=r'C:\Windows\msedgedriver.exe')
    driver = webdriver.Edge(service=service)
    driver.get(url)
    cards = get_cards(driver)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws['A1'] = 'Производитель'
    ws['B1'] = 'Описание'
    ws['C1'] = 'Ссылка на картинку'
    ws['D1'] = 'Цена'
    ws['E1'] = 'Оценка'
    ws['F1'] = 'Кол-во отзывов'
    for card in cards:
        data = parse(card)
        ws.append(data)

    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 44
    ws.column_dimensions['C'].width = 47

    wb.save("result.xlsx")




if __name__ == "__main__":
    main()
    print('Succesfully finished parsing.')

